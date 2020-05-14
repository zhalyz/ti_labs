#!/usr/bin/python3
import openpyxl


def init():
    raw_table = input('table 3x3 plz\n').split(' ')
    matrix = [[0 for x in range(3)] for y in range(3)]
    z = 0
    for x in range(3):
        for y in range(3):
            matrix[x][y] = int(raw_table[z])
            z += 1

    return matrix


def magic(table):

    wb = openpyxl.Workbook()
    wb.create_sheet(title='Первый лист', index=0)
    sheet = wb['Первый лист']

    tempa = 0
    tempb = 0
    shaga = 0
    shagb = 0

    an = 0
    bn = 0
    masA = [0 for i in range(3)]
    masB = [0 for i in range(3)]

    shag = 1

    nx = [0, 0, 0]
    ny = [0, 0, 0]

    print('k   A   B   x1  x2  x3    y1  y2  y3     v\     v/     e')
    a = an
    b = bn

    nx[an] += 1
    ny[bn] += 1

    for i in range(3):
        masA[i] = masA[i] + table[a][i]
        masB[i] = masB[i] + table[i][b]


    an = masB.index(max(masB))
    bn = masA.index(min(masA))

    if masB[an] == masB[a]:
        an = a

    if masA[bn] == masA[b]:
        bn = b

    kB = masA[bn]/(shag)
    kA = masB[an]/(shag)

    constKa = kA
    constKb = kB

    epselen = constKa - constKb

    sheet.cell(shag, 1).value = shag
    sheet.cell(shag, 2).value = a+1
    sheet.cell(shag, 3).value = b+1
    sheet.cell(shag, 4).value = masB[0]
    sheet.cell(shag, 5).value = masB[1]
    sheet.cell(shag, 6).value = masB[2]
    sheet.cell(shag, 7).value = masA[0]
    sheet.cell(shag, 8).value = masA[1]
    sheet.cell(shag, 9).value = masA[2]
    sheet.cell(shag, 10).value = str(masB[an]) + '/' + str(shag)
    sheet.cell(shag, 11).value = str(masA[bn]) + '/' + str(shag)
    sheet.cell(shag, 12).value = round(epselen, 4)
    print(shag, ' ', a+1, ' ', b+1, ' ', masB, '  ', masA, ' ',
          masB[an], '/', shag, '  ', masA[bn], '/', shag, ' ', round(epselen, 4))

    print((constKa+constKb)/2)
    shag += 1


    while abs(epselen) > 0.1:


        a = an
        b = bn

        nx[an] += 1
        ny[bn] += 1

        for i in range(3):
            masA[i] = masA[i] + table[a][i]
            masB[i] = masB[i] + table[i][b]


        an = masB.index(max(masB))
        bn = masA.index(min(masA))

        if masB[an] == masB[a]:
            an = a

        if masA[bn] == masA[b]:
            bn = b

        kB = masA[bn]/(shag)
        kA = masB[an]/(shag)

        if kB > constKb:
            tempb = masA[bn]
            shagb = shag
            constKb = kB

        if kA < constKa:
            tempa = masB[an]
            shaga = shag
            constKa = kA

        epselen = constKa - constKb

        sheet.cell(shag, 1).value = shag
        sheet.cell(shag, 2).value = a+1
        sheet.cell(shag, 3).value = b+1
        sheet.cell(shag, 4).value = masB[0]
        sheet.cell(shag, 5).value = masB[1]
        sheet.cell(shag, 6).value = masB[2]
        sheet.cell(shag, 7).value = masA[0]
        sheet.cell(shag, 8).value = masA[1]
        sheet.cell(shag, 9).value = masA[2]
        sheet.cell(shag, 10).value = str(masB[an]) + '/' + str(shag)
        sheet.cell(shag, 11).value = str(masA[bn]) + '/' + str(shag)
        sheet.cell(shag, 12).value = round(epselen, 4)

        print(shag, ' ', a+1, ' ', b+1, ' ', masB, '  ', masA, ' ',
              masB[an], '/', shag, '  ', masA[bn], '/', shag, ' ', round(epselen, 4))

        print((constKa+constKb)/2)

        shag += 1
    wb.save('example.xlsx')
    print(tempa, shaga, tempb, shagb)


if __name__ == "__main__":
    table = init()
    print(table)
    print()
    magic(table)

