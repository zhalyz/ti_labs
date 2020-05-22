#!/usr/bin/python3
import openpyxl

"""
-3 3/2 18/5 -18/50 -72/25
"""


def init():
    start = '-5 5/2 15 -3 -12'
    #table = input("input data").split()
    table = start.split()
    for i in range(len(table)):
        if '/' in table[i]:
            a, b = map(int, table[i].split('/'))

            table[i] = a/b
        else:
            table[i] = int(table[i])
    return table


def braunrobinson(xtable):

    ytable = tuple(zip(*xtable))
    # print(xtable)
    # print('-------------------------')
    shag = 1
    i, j = 0, 0
    xmas = [0 for i in range(len(ytable))]
    ymas = [0 for i in range(len(ytable))]
    x = [0 for i in range(len(xtable))]
    y = [0 for i in range(len(ytable))]
    xarray = []
    yarray = []
    while True:
        x[i] += 1
        y[j] += 1
        for arg in range(len(xtable)):
            xmas[arg] += ytable[j][arg]
            ymas[arg] += xtable[i][arg]

        xarray.append(max(xmas)/shag)
        yarray.append(min(ymas)/shag)

        epselen = min(xarray) - max(yarray)
        # print(shag)
        # print(i,j,xmas,ymas,max(xmas),'/',shag,min(ymas),'/',shag,epselen)

        if ymas[j] != min(ymas):
            j = ymas.index(min(ymas))

        if xmas[i] != max(xmas):
            i = xmas.index(max(xmas))
            # print(xmas,jpref,max(xmas))
        # print(i,j)
        # print()

        shag += 1
        if epselen <= 0.01:
            break
    # print(x,y)

    result = x.index(max(x)), y.index(max(y))
    return result


def aprox(table):
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Первый лист', index=0)
    sheet = wb['Первый лист']

    n = 2
    htable = [[0 for x in range(n+1)] for y in range(n+1)]

    print('\nN=', n)

    for i in range(n+1):
        for j in range(n+1):
            htable[i][j] = table[0] * ((i/n)**2) + table[1] * ((j/n)**2) + table[2] * (
                i/n) * (j/n) + table[3] * (i/n) + table[4] * (j/n)

    xarray = [0] * (n+1)
    yarray = [0] * (n+1)
    minx = [0] * (n+1)
    maxy = [0] * (n+1)
    rotate_htable = tuple(zip(*htable))

    for q in range(0, len(htable)):
        for e in range(0, len(htable[q])):
            print(round(htable[q][e], 3), end=' ')
            sheet.cell(n+q, e+1).value = round(htable[q][e], 3)
        print()

    for x in range(n+1):
        minx[x] = min(htable[x])
        maxy[x] = max(rotate_htable[x])
        xarray[x] = htable[x].index(minx[x])
        yarray[x] = rotate_htable[x].index(maxy[x])

    maxminx = max(minx)
    minmaxy = min(maxy)
    pres = 0
    if maxminx == minmaxy:
        print('Есть седловая точка:\nx= {0} y= {1} H= {2}'.format(
            round(((yarray[maxy.index(minmaxy)])/n), 4),
            round(((xarray[minx.index(maxminx)])/n), 4),
            round(maxminx, 4))
        )
        pres = maxminx
    else:
        print('Седловой точки нет, решение методом Брауна - Робинсон:')
        xres, yres = braunrobinson(htable)
        print('x=', round(xres/n, 4), 'y=', round(yres/n, 4),
              'H=', round(htable[xres][yres], 4))
        pres = htable[xres][yres]

    n += 1
    hy = 5

    while True:

        htable = [[0 for x in range(n+1)] for y in range(n+1)]
        print('\nN=', n)

        for i in range(n+1):
            for j in range(n+1):
                htable[i][j] = table[0] * ((i/n)**2) + table[1] * ((j/n)**2) + table[2] * (
                    i/n) * (j/n) + table[3] * (i/n) + table[4] * (j/n)

        xarray = [0] * (n+1)
        yarray = [0] * (n+1)
        minx = [0] * (n+1)
        maxy = [0] * (n+1)
        rotate_htable = tuple(zip(*htable))

        for q in range(0, len(htable)):
            for e in range(0, len(htable[q])):
                print(round(htable[q][e], 3), end=' ')
                sheet.cell(hy, e+1).value = round(htable[q][e], 3)
            print()
            hy += 1

        for x in range(n+1):
            minx[x] = min(htable[x])
            maxy[x] = max(rotate_htable[x])
            xarray[x] = htable[x].index(minx[x])
            yarray[x] = rotate_htable[x].index(maxy[x])

        maxminx = max(minx)
        minmaxy = min(maxy)
        if maxminx == minmaxy:
            print('Есть седловая точка:\nx= {0} y= {1} H= {2}'.format(
                round(((yarray[maxy.index(minmaxy)])/n), 4),
                round(((xarray[minx.index(maxminx)])/n), 4),
                round(maxminx, 4))
            )
            hres = maxminx

        else:
            print('Седловой точки нет, решение методом Брауна - Робинсон:')
            xres, yres = braunrobinson(htable)
            print('x=', round(xres/n, 4), 'y=', round(yres/n, 4),
                  'H=', round(htable[xres][yres], 4))
            hres = htable[xres][yres]

        if abs(pres - hres) <= 0.001:
            wb.save('example.xlsx')
            break
        n += 1
        pres = hres


if __name__ == "__main__":
    table = init()
    print(table)
    aprox(table)
